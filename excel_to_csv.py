import sys
import csv
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QFileDialog, QVBoxLayout,
    QWidget, QProgressBar, QLabel, QLineEdit, QMessageBox
)
from PyQt5.QtCore import Qt
import os

class ExcelToCSVConverter(QMainWindow):
    def __init__(self):
        super().__init__()

        self.initUI()
        self.file_path = None
        self.destination_directory = None

    def initUI(self):
        # Set up the main window
        self.setWindowTitle('Excel to CSV Converter')
        self.setGeometry(100, 100, 500, 250)

        # Create layout
        layout = QVBoxLayout()

        # Create input field for browsing Excel file
        self.file_input = QLineEdit(self)
        browse_file_button = QPushButton('Browse Excel File', self)
        browse_file_button.clicked.connect(self.browse_file)

        # Create input field for browsing destination directory
        self.destination_input = QLineEdit(self)
        browse_dest_button = QPushButton('Select Destination Directory', self)
        browse_dest_button.clicked.connect(self.select_destination_directory)

        # Create and add a progress bar
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setValue(0)

        # Create a label to show the status
        self.status_label = QLabel('', self)

        # Create a close button
        close_button = QPushButton('Close', self)
        close_button.clicked.connect(self.close_application)

        # Create process button
        process_button = QPushButton('Convert', self)
        process_button.clicked.connect(self.convert_excel_to_csv)

        # Add widgets to layout
        layout.addWidget(QLabel('Select Excel File:'))
        layout.addWidget(self.file_input)
        layout.addWidget(browse_file_button)
        layout.addWidget(QLabel('Select Destination Directory:'))
        layout.addWidget(self.destination_input)
        layout.addWidget(browse_dest_button)
        layout.addWidget(self.progress_bar)
        layout.addWidget(process_button)
        layout.addWidget(self.status_label)
        layout.addWidget(close_button)

        # Set central widget
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def browse_file(self):
        # Open file dialog to select an Excel file
        file_dialog = QFileDialog(self)
        self.file_path, _ = file_dialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if self.file_path:
            self.file_input.setText(self.file_path)

    def select_destination_directory(self):
        # Open directory dialog to select a destination directory
        directory_dialog = QFileDialog(self)
        self.destination_directory = directory_dialog.getExistingDirectory(self, 'Select Destination Directory')
        if self.destination_directory:
            self.destination_input.setText(self.destination_directory)

    def convert_excel_to_csv(self):
        if not self.file_path:
            QMessageBox.warning(self, 'Warning', 'Please select an Excel file first.')
            return

        if not self.destination_directory:
            QMessageBox.warning(self, 'Warning', 'Please select a destination directory first.')
            return

        try:
            workbook = load_workbook(self.file_path)
            sheet_names = workbook.sheetnames
            total_sheets = len(sheet_names)
            self.progress_bar.setMaximum(total_sheets)

            for idx, sheet_name in enumerate(sheet_names):
                sheet = workbook[sheet_name]

                # Remove the first 7 rows
                for _ in range(7):
                    sheet.delete_rows(1)

                # Save the sheet to a CSV file
                csv_filename = os.path.join(self.destination_directory, f"{sheet_name}.csv")
                with open(csv_filename, 'w', newline='', encoding='utf-8') as csv_file:
                    writer = csv.writer(csv_file)
                    for row in sheet.iter_rows(values_only=True):
                        writer.writerow(row)

                # Update progress bar
                self.progress_bar.setValue(idx + 1)
                QApplication.processEvents()

            self.status_label.setText('Conversion completed successfully!')
            QMessageBox.information(self, 'Success', 'Conversion completed successfully!')

        except Exception as e:
            QMessageBox.critical(self, 'Error', f'An error occurred: {e}')

    def close_application(self):
        self.close()

def main():
    app = QApplication(sys.argv)
    converter = ExcelToCSVConverter()
    converter.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
