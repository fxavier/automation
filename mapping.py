import sys
import requests
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QFileDialog, QVBoxLayout,
    QWidget, QProgressBar, QLabel, QLineEdit, QFormLayout, QMessageBox
)
from PyQt5.QtCore import Qt

class DHIS2DataMapper(QMainWindow):
    def __init__(self):
        super().__init__()

        self.initUI()
        self.file_path = None
        self.username = ''
        self.password = ''
        self.base_url = ''

    def initUI(self):
        # Set up the main window
        self.setWindowTitle('DHIS2 Data Element Mapper')
        self.setGeometry(100, 100, 500, 300)

        # Create layout
        main_layout = QVBoxLayout()

        # Create form layout for DHIS2 credentials
        form_layout = QFormLayout()

        # Create input fields for DHIS2 credentials
        self.username_input = QLineEdit(self)
        self.password_input = QLineEdit(self)
        self.password_input.setEchoMode(QLineEdit.Password)  # Mask password input
        self.base_url_input = QLineEdit(self)

        # Add input fields to the form layout
        form_layout.addRow('DHIS2 Username:', self.username_input)
        form_layout.addRow('DHIS2 Password:', self.password_input)
        form_layout.addRow('DHIS2 Base URL:', self.base_url_input)

        # Add form layout to the main layout
        main_layout.addLayout(form_layout)

        # Create and add the browse button
        self.browse_button = QPushButton('Browse Excel File', self)
        self.browse_button.clicked.connect(self.browse_file)
        main_layout.addWidget(self.browse_button)

        # Create and add the process button
        self.process_button = QPushButton('Process', self)
        self.process_button.setEnabled(False)  # Initially disabled until a file is selected
        self.process_button.clicked.connect(self.map_data_elements)
        main_layout.addWidget(self.process_button)

        # Create and add the progress bar
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setValue(0)
        main_layout.addWidget(self.progress_bar)

        # Create and add the status label
        self.status_label = QLabel('', self)
        main_layout.addWidget(self.status_label)

        # Create and add the close button
        self.close_button = QPushButton('Close', self)
        self.close_button.clicked.connect(self.close_application)
        main_layout.addWidget(self.close_button)

        # Set central widget
        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

    def browse_file(self):
        # Check if credentials are filled
        self.username = self.username_input.text()
        self.password = self.password_input.text()
        self.base_url = self.base_url_input.text() + '/api/'

        if not self.username or not self.password or not self.base_url:
            QMessageBox.warning(self, 'Input Error', 'Please fill in all DHIS2 credentials before browsing for a file.')
            return

        # Open file dialog to select an Excel file
        file_dialog = QFileDialog(self)
        self.file_path, _ = file_dialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx)')

        if self.file_path:
            self.status_label.setText(f'File Selected: {self.file_path}')
            self.process_button.setEnabled(True)  # Enable the Process button

    def map_data_elements(self):
        if not self.file_path:
            self.status_label.setText('No file selected.')
            return

        # Load the Excel file using openpyxl
        workbook = load_workbook(self.file_path)
        sheets = workbook.sheetnames

        total_sheets = len(sheets)
        self.progress_bar.setMaximum(total_sheets)

        session = requests.Session()
        session.auth = (self.username, self.password)

        # Process each sheet
        for sheet_index, sheet_name in enumerate(sheets):
            sheet = workbook[sheet_name]
            # Read all non-empty cells in row 8
            data_elements = {cell.column: cell.value for cell in sheet[8] if cell.value}

            # Search data elements
            results = {}
            for col, data_element_name in data_elements.items():
                if data_element_name:  # Ensure the data element name is not None or empty
                    results[col] = self.search_data_element(data_element_name, session)
                else:
                    results[col] = 'Invalid Header'

            # Write results back to the sheet in row 9
            for col, result in results.items():
                sheet.cell(row=9, column=col, value=result)

            # Update progress bar
            self.progress_bar.setValue(sheet_index + 1)

        # Save the updated Excel file
        output_file = self.file_path.replace('.xlsx', '_Updated.xlsx')
        workbook.save(output_file)
        self.status_label.setText(f'Updated file saved to {output_file}')

    def search_data_element(self, data_element_name, session):
        try:
            response = session.get(self.base_url + 'dataElements.json',
                                   params={'filter': f'name:ilike:{data_element_name}', 'fields': 'id,name', 'paging': 'false'})
            response.raise_for_status()  # Raises an HTTPError for bad responses
            data_elements = response.json().get('dataElements', [])
            if data_elements:
                return data_elements[0]['id']  # Assuming the first match is the correct one
            else:
                print(f"No data element found for: {data_element_name}")
                return None
        except requests.exceptions.HTTPError as http_err:
            print(f"HTTP error occurred: {http_err}")  # Improved logging
            return 'HTTP Error'
        except requests.exceptions.RequestException as req_err:
            print(f"Request error occurred: {req_err}")  # Improved logging
            return 'Request Error'
        except Exception as e:
            print(f"An error occurred: {e}")  # General exception logging
            return 'Error'

    def close_application(self):
        self.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = DHIS2DataMapper()
    window.show()
    sys.exit(app.exec_())
