import sys
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QFileDialog, QMessageBox, QProgressBar
)
import openpyxl

class ExcelMerger(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Excel File Merger")

        # Layouts
        layout = QVBoxLayout()
        layout_files = QHBoxLayout()

        # File selection
        label_files = QLabel("Select Excel Files:")
        self.entry_files = QLineEdit(self)
        btn_files = QPushButton("Browse")
        btn_files.clicked.connect(self.select_files)

        layout_files.addWidget(label_files)
        layout_files.addWidget(self.entry_files)
        layout_files.addWidget(btn_files)

        # Progress bar
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setValue(0)

        # Merge button
        btn_merge = QPushButton("Merge Files")
        btn_merge.setStyleSheet("background-color: green; color: white;")
        btn_merge.clicked.connect(self.merge_files)

        # Close button
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(self.close_application)

        # Add layouts to main layout
        layout.addLayout(layout_files)
        layout.addWidget(self.progress_bar)
        layout.addWidget(btn_merge)
        layout.addWidget(btn_close)

        self.setLayout(layout)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Excel Files", "", "Excel files (*.xlsx *.xls)")
        if files:
            self.entry_files.setText(";".join(files))

    def merge_files(self):
        file_paths = self.entry_files.text().split(';')

        if not file_paths or len(file_paths) < 2:
            QMessageBox.warning(self, "Warning", "Please select at least two Excel files.")
            return

        try:
            # Load the first workbook
            wb_dst = openpyxl.load_workbook(file_paths[0])

            total_steps = len(file_paths) - 1  # Total number of files to process
            self.progress_bar.setMaximum(total_steps)

            for idx, file_path in enumerate(file_paths[1:], start=1):
                wb_src = openpyxl.load_workbook(file_path)

                for ws_src in wb_src.worksheets:
                    # Check if sheet exists in destination workbook
                    if ws_src.title in wb_dst.sheetnames:
                        ws_dst = wb_dst[ws_src.title]
                    else:
                        ws_dst = wb_dst.create_sheet(ws_src.title)

                    # Copy rows from the 8th row of the source sheet
                    for row in ws_src.iter_rows(min_row=8, values_only=True):
                        if any(row):  # Check if the row is not entirely empty
                            ws_dst.append(row)

                # Update progress bar
                self.progress_bar.setValue(idx)
                QApplication.processEvents()  # Process UI events

            # Save the merged workbook
            merged_file_path, _ = QFileDialog.getSaveFileName(self, "Save Merged File", "", "Excel files (*.xlsx)")
            if merged_file_path:
                wb_dst.save(merged_file_path)
                QMessageBox.information(self, "Success", "Files merged successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {e}")

    def close_application(self):
        self.close()

def main():
    app = QApplication(sys.argv)
    merger = ExcelMerger()
    merger.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
