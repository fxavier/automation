import sys
import csv
import os
import requests
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QFileDialog, QMessageBox, QProgressBar, QFormLayout, QComboBox
)
from PyQt5.QtCore import Qt, pyqtSignal
import pandas as pd 

class MainApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Excel Utility Tool')
        self.setGeometry(100, 100, 600, 400)

        layout = QVBoxLayout()

        btn_excel_merger = QPushButton('Excel File Merger', self)
        btn_excel_merger.clicked.connect(self.open_excel_merger)
        layout.addWidget(btn_excel_merger)

        btn_excel_to_csv = QPushButton('Excel to CSV Converter', self)
        btn_excel_to_csv.clicked.connect(self.open_excel_to_csv)
        layout.addWidget(btn_excel_to_csv)

        btn_dhis2_mapper = QPushButton('DHIS2 Data Element Mapper', self)
        btn_dhis2_mapper.clicked.connect(self.open_dhis2_mapper)
        layout.addWidget(btn_dhis2_mapper)

        btn_period_selector = QPushButton('DHIS2 Period Selector', self)
        btn_period_selector.clicked.connect(self.open_period_selector)
        layout.addWidget(btn_period_selector)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def open_excel_merger(self):
        self.excel_merger = ExcelMerger()
        self.excel_merger.show()

    def open_excel_to_csv(self):
        self.excel_to_csv = ExcelToCSVConverter()
        self.excel_to_csv.show()

    def open_dhis2_mapper(self):
        self.dhis2_mapper = DHIS2DataMapper()
        self.dhis2_mapper.show()

    def open_period_selector(self):
        self.period_selector = DHIS2PeriodSelector(self)
        self.period_selector.show()

class ExcelMerger(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Excel File Merger")

        # Layouts
        layout = QVBoxLayout()
        layout_files = QHBoxLayout()
        layout_files_two = QHBoxLayout()

        # File selection for multiple files
        label_files = QLabel("Select Excel Files (Multiple):")
        self.entry_files = QLineEdit(self)
        btn_files = QPushButton("Browse")
        btn_files.clicked.connect(self.select_files)

        layout_files.addWidget(label_files)
        layout_files.addWidget(self.entry_files)
        layout_files.addWidget(btn_files)

        # File selection for two files
        label_files_two = QLabel("Select Two Excel Files:")
        self.entry_files_two = QLineEdit(self)
        btn_files_two = QPushButton("Browse")
        btn_files_two.clicked.connect(self.select_two_files)

        layout_files_two.addWidget(label_files_two)
        layout_files_two.addWidget(self.entry_files_two)
        layout_files_two.addWidget(btn_files_two)

        # Progress bars
        self.progress_bar_multiple = QProgressBar(self)
        self.progress_bar_multiple.setMinimum(0)
        self.progress_bar_multiple.setValue(0)

        self.progress_bar_two_files = QProgressBar(self)
        self.progress_bar_two_files.setMinimum(0)
        self.progress_bar_two_files.setValue(0)

        # Merge buttons
        btn_merge = QPushButton("Merge Multiple Files")
        btn_merge.setStyleSheet("background-color: green; color: white;")
        btn_merge.clicked.connect(self.merge_files)

        btn_merge_two = QPushButton("Mapping Files")
        btn_merge_two.setStyleSheet("background-color: blue; color: white;")
        btn_merge_two.clicked.connect(self.merge_two_files)

        # Close button
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(self.close_application)

        # Add layouts to main layout
        layout.addLayout(layout_files)
        layout.addWidget(self.progress_bar_multiple)
        layout.addLayout(layout_files_two)
        layout.addWidget(self.progress_bar_two_files)
        layout.addWidget(btn_merge)
        layout.addWidget(btn_merge_two)
        layout.addWidget(btn_close)

        self.setLayout(layout)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Excel Files", "", "Excel files (*.xlsx *.xls)")
        if files:
            self.entry_files.setText(";".join(files))

    def select_two_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Two Excel Files", "", "Excel files (*.xlsx *.xls)")
        if len(files) == 2:
            self.entry_files_two.setText(";".join(files))
        else:
            QMessageBox.warning(self, "Warning", "Please select exactly two Excel files.")

    def merge_files(self):
        file_paths = self.entry_files.text().split(';')

        if not file_paths or len(file_paths) < 2:
            QMessageBox.warning(self, "Warning", "Please select at least two Excel files.")
            return

        try:
            wb_dst = load_workbook(file_paths[0])

            total_steps = len(file_paths) - 1
            self.progress_bar_multiple.setMaximum(total_steps)

            for idx, file_path in enumerate(file_paths[1:], start=1):
                wb_src = load_workbook(file_path)

                for ws_src in wb_src.worksheets:
                    if ws_src.title in wb_dst.sheetnames:
                        ws_dst = wb_dst[ws_src.title]
                    else:
                        ws_dst = wb_dst.create_sheet(ws_src.title)

                    for row in ws_src.iter_rows(min_row=8, values_only=True):
                        if any(row):
                            ws_dst.append(row)

                self.progress_bar_multiple.setValue(idx)
                QApplication.processEvents()

            merged_file_path, _ = QFileDialog.getSaveFileName(self, "Save Merged File", "", "Excel files (*.xlsx)")
            if merged_file_path:
                wb_dst.save(merged_file_path)
                QMessageBox.information(self, "Success", "Files merged successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {e}")

    def merge_two_files(self):
        file_paths = self.entry_files_two.text().split(';')

        if not file_paths or len(file_paths) != 2:
            QMessageBox.warning(self, "Warning", "Please select exactly two Excel files.")
            return

        try:
            wb_dst = load_workbook(file_paths[0])
            wb_src = load_workbook(file_paths[1])

            total_sheets = len(wb_src.worksheets)
            self.progress_bar_two_files.setMaximum(total_sheets)

            for idx, ws_src in enumerate(wb_src.worksheets, start=1):
                if ws_src.title in wb_dst.sheetnames:
                    ws_dst = wb_dst[ws_src.title]
                else:
                    ws_dst = wb_dst.create_sheet(ws_src.title)

                # Copy rows from the 8th row of the source sheet
                for row in ws_src.iter_rows(min_row=8, values_only=True):
                    if any(row):  # Check if the row is not entirely empty
                        ws_dst.append(row)

                # Move the last non-empty line to line 8
                self.move_last_non_empty_line(ws_dst)

                # Update progress bar for merging two files
                self.progress_bar_two_files.setValue(idx)
                QApplication.processEvents()

            merged_file_path, _ = QFileDialog.getSaveFileName(self, "Save Merged File", "", "Excel files (*.xlsx)")
            if merged_file_path:
                wb_dst.save(merged_file_path)
                QMessageBox.information(self, "Success", "Files merged successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {e}")

    def move_last_non_empty_line(self, worksheet):
        """Move the last non-empty line to line 8, shifting down the remaining data."""
        max_row = worksheet.max_row
        last_non_empty_row = max_row

        # Find the last non-empty row
        for row in range(max_row, 0, -1):
            if any(worksheet.cell(row=row, column=col).value not in (None, '', ' ') for col in range(1, worksheet.max_column + 1)):
                last_non_empty_row = row
                break

        # Extract the values from the last non-empty row
        last_row_values = [worksheet.cell(row=last_non_empty_row, column=col).value for col in range(1, worksheet.max_column + 1)]

        # Delete the last non-empty row
        worksheet.delete_rows(last_non_empty_row)

        # Shift all rows from line 8 down by one row
        worksheet.insert_rows(8)

        # Paste the extracted values into line 8
        for col, value in enumerate(last_row_values, start=1):
            worksheet.cell(row=8, column=col, value=value)

    def close_application(self):
        self.close()


class ExcelToCSVConverter(QMainWindow):
    def __init__(self):
        super().__init__()
        self.file_path = None
        self.destination_directory = None
        self.period_code = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Excel to CSV Converter')
        self.setGeometry(100, 100, 500, 250)

        layout = QVBoxLayout()

        # File input and destination directory selection
        layout.addWidget(QLabel('Select Excel File:'))
        self.file_input = QLineEdit(self)
        layout.addWidget(self.file_input)
        browse_file_button = QPushButton('Browse Excel File', self)
        browse_file_button.clicked.connect(self.browse_file)
        layout.addWidget(browse_file_button)

        layout.addWidget(QLabel('Select Destination Directory:'))
        self.destination_input = QLineEdit(self)
        layout.addWidget(self.destination_input)
        browse_dest_button = QPushButton('Select Destination Directory', self)
        browse_dest_button.clicked.connect(self.select_destination_directory)
        layout.addWidget(browse_dest_button)

        # Progress bar and status label
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        self.status_label = QLabel('', self)
        layout.addWidget(self.status_label)

        # Action buttons
        self.period_selector_button = QPushButton('Select Period', self)
        self.period_selector_button.clicked.connect(self.open_period_selector)
        layout.addWidget(self.period_selector_button)

        process_button = QPushButton('Convert', self)
        process_button.clicked.connect(self.convert_excel_to_csv)
        layout.addWidget(process_button)

        close_button = QPushButton('Close', self)
        close_button.clicked.connect(self.close_application)
        layout.addWidget(close_button)

        # Set central widget
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def browse_file(self):
        self.file_path, _ = QFileDialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if self.file_path:
            self.file_input.setText(self.file_path)

    def select_destination_directory(self):
        self.destination_directory = QFileDialog.getExistingDirectory(self, 'Select Destination Directory')
        if self.destination_directory:
            self.destination_input.setText(self.destination_directory)

    def open_period_selector(self):
        self.period_selector = DHIS2PeriodSelector(self)
        self.period_selector.period_selected.connect(self.set_period_code)
        self.period_selector.show()

    def set_period_code(self, period_code):
        self.period_code = period_code
        QMessageBox.information(self, 'Period Code Set', f'Period code set to: {period_code}')

    def convert_excel_to_csv(self):
        if not self.file_path or not self.destination_directory or not self.period_code:
            QMessageBox.warning(self, 'Warning', 'Please ensure all fields are filled.')
            return

        try:
            workbook = load_workbook(self.file_path)
            sheet_names = workbook.sheetnames
            self.progress_bar.setMaximum(len(sheet_names))

            for idx, sheet_name in enumerate(sheet_names):
                if sheet_name == "PrEP Extra Dissag":
                    continue

                self.process_sheet(workbook[sheet_name], sheet_name)
                self.progress_bar.setValue(idx + 1)
                QApplication.processEvents()

            # Set progress bar to 100% after processing all sheets
            self.progress_bar.setValue(len(sheet_names))  # Ensure it reaches 100%

            # Call additional processing steps
            self.unpivot_columns()
            transformed_files = os.listdir(os.path.join(os.path.dirname(__file__), 'Transformed'))
            final_df = self.concat_final_files(transformed_files)
            final_df = self.prepare_final_file(final_df)
            final_df.to_csv(os.path.join(os.path.dirname(__file__), 'Merged', 'file_import.csv'), index=False)

            # Convert dataelements with dots
            df = pd.read_csv(os.path.join(os.path.dirname(__file__), 'Merged', 'file_import.csv'))
            df[['dataelement', 'categoryoptioncombo']] = df['dataelement'].str.split('.', expand=True)
            df.to_csv(os.path.join(os.path.dirname(__file__), 'Merged', 'file_to_import.csv'), index=False)

            self.status_label.setText('Conversion and merging completed successfully!')
            QMessageBox.information(self, 'Success', 'Conversion and merging completed successfully!')

        except Exception as e:
            QMessageBox.critical(self, 'Error', f'An error occurred: {e}')

    def process_sheet(self, sheet, sheet_name):
        for _ in range(7):
            sheet.delete_rows(1)

        rows = list(sheet.iter_rows(values_only=True))
        headers = self.prepare_headers(rows[0])
        non_empty_indices = [i for i, header in enumerate(headers) if header and header.strip()]
        filtered_headers = [headers[i] for i in non_empty_indices]

        csv_filename = os.path.join(self.destination_directory, f"{sheet_name}.csv")
        with open(csv_filename, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(filtered_headers)

            for row in rows[1:]:
                row = list(row)
                row[0], row[2] = row[2], self.period_code
                filtered_row = [row[i] for i in non_empty_indices]
                writer.writerow(filtered_row)

        self.merge_with_orgunits(csv_filename)

    def prepare_headers(self, headers):
        headers = list(headers)
        headers[0], headers[2] = 'Datim_Code', 'period'
        return headers

    def merge_with_orgunits(self, csv_filename):
        try:
            df_csv = pd.read_csv(csv_filename)
            orgunits_file = os.path.join(os.path.dirname(__file__), 'orgunits', 'orgunits.csv')
            df_orgunits = pd.read_csv(orgunits_file)

            if 'orgunit' not in df_orgunits.columns:
                raise KeyError("The 'orgunit' column is missing in the orgunits file.")

            merged_df = pd.merge(df_csv, df_orgunits, on='Datim_Code', how='inner')
            merged_df.to_csv(csv_filename, index=False)

        except KeyError as e:
            QMessageBox.critical(self, 'Error', f'Key error: {e}')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'An error occurred while merging with orgunits: {e}')

    def unpivot_columns(self):
        initial_files_dir = self.destination_directory
        transformed_dir = os.path.join(os.path.dirname(__file__), 'Transformed')
        os.makedirs(transformed_dir, exist_ok=True)

        for file in os.listdir(initial_files_dir):
            if file.endswith('.csv'):
                df = pd.read_csv(os.path.join(initial_files_dir, file))

                # Ensure required columns exist
                required_columns = ['orgunit', 'Province', 'District', 'Health Facility']
                missing_columns = [col for col in required_columns if col not in df.columns]

                if missing_columns:
                    QMessageBox.warning(self, 'Warning', f"Missing columns in {file}: {', '.join(missing_columns)}")
                    continue  # Skip processing this file

                column_excluded = ['Datim_Code', 'period', 'Province', 'District', 'Health Facility', 'orgunit']
                columns_to_unpivot = df.loc[:, ~df.columns.isin(column_excluded)]
                column_list = [x for x in columns_to_unpivot]

                try:
                    df = df.melt(
                        id_vars=['orgunit', 'period', 'Province', 'District', 'Health Facility'],
                        var_name='dataelement',
                        value_vars=column_list
                    )
                    df.to_csv(os.path.join(transformed_dir, file), index=False)
                except Exception as e:
                    QMessageBox.critical(self, 'Error', f"An error occurred while unpivoting {file}: {e}")

    def concat_final_files(self, transformed_files):
        df_list = []
        transformed_dir = os.path.join(os.path.dirname(__file__), 'Transformed')
        for transformed_file in transformed_files:
            try:
                df = pd.read_csv(os.path.join(transformed_dir, transformed_file))
                df_list.append(df)
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'An error occurred while concatenating files: {e}')
        if not df_list:
            raise ValueError("No valid files to concatenate.")
        final_df = pd.concat(df_list).drop_duplicates().reset_index(drop=True)
        final_df.to_csv(os.path.join(os.path.dirname(__file__), 'Final_Files', 'concatenated_file.csv'), index=False)
        return final_df

    def prepare_final_file(self, df):
        df = df[['dataelement', 'period', 'orgunit', 'value']]
        df.columns = ['dataelement', 'period', 'orgunit', 'value']
        df = df.assign(categoryoptioncombo=[''] * df.shape[0])
        df = df.assign(attributeoptioncombo=[''] * df.shape[0])
        return df.iloc[:, [0, 1, 2, 4, 5, 3]]

    def close_application(self):
        self.close()


class DHIS2DataMapper(QMainWindow):
    def __init__(self):
        super().__init__()

        self.initUI()
        self.file_path = None
        self.username = ''
        self.password = ''
        self.base_url = ''

    def initUI(self):
        # Set up the main window
        self.setWindowTitle('DHIS2 Data Element Mapper')
        self.setGeometry(100, 100, 500, 300)

        # Create layout
        main_layout = QVBoxLayout()

        # Create form layout for DHIS2 credentials
        form_layout = QFormLayout()

        # Create input fields for DHIS2 credentials
        self.username_input = QLineEdit(self)
        self.password_input = QLineEdit(self)
        self.password_input.setEchoMode(QLineEdit.Password)  # Mask password input
        self.base_url_input = QLineEdit(self)

        # Add input fields to the form layout
        form_layout.addRow('DHIS2 Username:', self.username_input)
        form_layout.addRow('DHIS2 Password:', self.password_input)
        form_layout.addRow('DHIS2 Base URL:', self.base_url_input)

        # Add form layout to the main layout
        main_layout.addLayout(form_layout)

        # Create and add the browse button
        self.browse_button = QPushButton('Browse Excel File', self)
        self.browse_button.clicked.connect(self.browse_file)
        main_layout.addWidget(self.browse_button)

        # Create and add the process button
        self.process_button = QPushButton('Process', self)
        self.process_button.setEnabled(False)  # Initially disabled until a file is selected
        self.process_button.clicked.connect(self.map_data_elements)
        main_layout.addWidget(self.process_button)

        # Create and add the progress bar
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setValue(0)
        main_layout.addWidget(self.progress_bar)

        # Create and add the status label
        self.status_label = QLabel('', self)
        main_layout.addWidget(self.status_label)

        # Create and add the close button
        self.close_button = QPushButton('Close', self)
        self.close_button.clicked.connect(self.close_application)
        main_layout.addWidget(self.close_button)

        # Set central widget
        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

    def browse_file(self):
        # Check if credentials are filled
        self.username = self.username_input.text()
        self.password = self.password_input.text()
        self.base_url = self.base_url_input.text() + '/api/'

        if not self.username or not self.password or not self.base_url:
            QMessageBox.warning(self, 'Input Error', 'Please fill in all DHIS2 credentials before browsing for a file.')
            return

        # Open file dialog to select an Excel file
        file_dialog = QFileDialog(self)
        self.file_path, _ = file_dialog.getOpenFileName(self, 'Open Excel File', '', 'Excel Files (*.xlsx)')

        if self.file_path:
            self.status_label.setText(f'File Selected: {self.file_path}')
            self.process_button.setEnabled(True)  # Enable the Process button

    def map_data_elements(self):
        if not self.file_path:
            self.status_label.setText('No file selected.')
            return

        # Load the Excel file using openpyxl
        workbook = load_workbook(self.file_path)
        sheets = workbook.sheetnames

        total_sheets = len(sheets)
        self.progress_bar.setMaximum(total_sheets)

        session = requests.Session()
        session.auth = (self.username, self.password)

        # Process each sheet
        for sheet_index, sheet_name in enumerate(sheets):
            sheet = workbook[sheet_name]
            # Read all non-empty cells in row 8
            data_elements = {cell.column: cell.value for cell in sheet[8] if cell.value}

            # Search data elements
            results = {}
            for col, data_element_name in data_elements.items():
                if data_element_name:  # Ensure the data element name is not None or empty
                    results[col] = self.search_data_element(data_element_name, session)
                else:
                    results[col] = 'Invalid Header'

            # Write results back to the sheet in row 9
            for col, result in results.items():
                sheet.cell(row=9, column=col, value=result)

            # Update progress bar
            self.progress_bar.setValue(sheet_index + 1)

        # Save the updated Excel file
        output_file = self.file_path.replace('.xlsx', '_Updated.xlsx')
        workbook.save(output_file)
        self.status_label.setText(f'Updated file saved to {output_file}')

    def search_data_element(self, data_element_name, session):
        try:
            response = session.get(self.base_url + 'dataElements.json',
                                   params={'filter': f'name:ilike:{data_element_name}', 'fields': 'id,name', 'paging': 'false'})
            response.raise_for_status()  # Raises an HTTPError for bad responses
            data_elements = response.json().get('dataElements', [])
            if data_elements:
                return data_elements[0]['id']  # Assuming the first match is the correct one
            else:
                print(f"No data element found for: {data_element_name}")
                return None
        except requests.exceptions.HTTPError as http_err:
            print(f"HTTP error occurred: {http_err}")  # Improved logging
            return 'HTTP Error'
        except requests.exceptions.RequestException as req_err:
            print(f"Request error occurred: {req_err}")  # Improved logging
            return 'Request Error'
        except Exception as e:
            print(f"An error occurred: {e}")  # General exception logging
            return 'Error'

    def close_application(self):
        self.close()

class DHIS2PeriodSelector(QMainWindow):
    period_selected = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()

    def initUI(self):
        self.setWindowTitle('DHIS2 Period Selector')
        self.setGeometry(100, 100, 400, 200)

        main_layout = QVBoxLayout()

        # Period Type Combobox
        self.period_type_label = QLabel('Period Type:', self)
        self.period_type_combo = QComboBox(self)
        self.period_type_combo.addItems(['Monthly', 'Quarterly', 'Semi-annual', 'Annual'])
        main_layout.addWidget(self.period_type_label)
        main_layout.addWidget(self.period_type_combo)

        # Year Combobox
        self.year_label = QLabel('Year:', self)
        self.year_combo = QComboBox(self)
        self.year_combo.addItems([str(year) for year in range(2019, 2027)])
        main_layout.addWidget(self.year_label)
        main_layout.addWidget(self.year_combo)

        # Period Combobox
        self.period_label = QLabel('Period:', self)
        self.period_combo = QComboBox(self)
        self.update_periods()
        main_layout.addWidget(self.period_label)
        main_layout.addWidget(self.period_combo)

        # Update periods when period type or year changes
        self.period_type_combo.currentIndexChanged.connect(self.update_periods)
        self.year_combo.currentIndexChanged.connect(self.update_periods)

        # Submit Button
        self.submit_button = QPushButton('Submit', self)
        self.submit_button.clicked.connect(self.submit)
        main_layout.addWidget(self.submit_button)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

    def update_periods(self):
        period_type = self.period_type_combo.currentText()
        year = int(self.year_combo.currentText())
        self.period_combo.clear()

        if period_type == 'Monthly':
            months = [
                'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December'
            ]
            self.period_combo.addItems([f'{month}-{year}' for month in months])
        elif period_type == 'Quarterly':
            self.period_combo.addItems([
                f'January to March {year}', 
                f'April to June {year}', 
                f'July to October {year}', 
                f'September to December {year}'
            ])
        elif period_type == 'Semi-annual':
            self.period_combo.addItems([
                f'April - September {year}',
                f'October - March {year + 1}'
            ])
        elif period_type == 'Annual':
            self.period_combo.addItems([
                f'October {year -1 } - September {year}'
            ])

    def submit(self):
        period_type = self.period_type_combo.currentText()
        year = int(self.year_combo.currentText())
        period = self.period_combo.currentText()

        # Determine the period code
        if period_type == 'Monthly':
            month = period.split('-')[0]
            month_number = {
                'January': '01', 'February': '02', 'March': '03', 'April': '04',
                'May': '05', 'June': '06', 'July': '07', 'August': '08',
                'September': '09', 'October': '10', 'November': '11', 'December': '12'
            }[month]
            period_code = f'{year}{month_number}'
        elif period_type == 'Quarterly':
            quarter = period.split(' ')[0]
            quarter_number = {
                'January': 'Q1', 'April': 'Q2', 'July': 'Q3', 'September': 'Q4'
            }[quarter]
            period_code = f'{year}{quarter_number}'
        elif period_type == 'Semi-annual':
            semi_annual = period.split(' ')[0]
            semi_annual_code = {
                'April': 'AprilS1', 'October': 'AprilS2'
            }[semi_annual]
            period_code = f'{year}{semi_annual_code}'
        elif period_type == 'Annual':
            period_code = f'{year-1}Oct'

        # Emit the period code
        self.period_selected.emit(period_code)
        self.close()

def main():
    app = QApplication(sys.argv)
    main_app = MainApp()
    main_app.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()